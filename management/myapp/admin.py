from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.utils.http import urlencode
from import_export.admin import ExportActionMixin
from django import forms
from django.db.models import Sum
import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from django.http import HttpResponse
from myapp.models import (
    Client,
    Expenses,
    Vendor,
    Store,
    StockItems,
    StockInward,
    StockOutward,
    StockAdjustment,
)
from django.urls import path, reverse


class ClientAdmin(admin.ModelAdmin):
    list_display = ("name", "email", "phone", "city", "state")
    search_fields = list_display
    list_per_page = 25


    def export_to_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="clients.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Clients"

        # Write column headers
        headers = ["Name", "Email", "Phone", "City", "State"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_num, client in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=client.name)
            worksheet.cell(row=row_num, column=2, value=client.email)
            worksheet.cell(row=row_num, column=3, value=client.phone)
            worksheet.cell(row=row_num, column=4, value=client.city)
            worksheet.cell(row=row_num, column=5, value=client.state)

        # Auto-size columns
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save workbook to response
        workbook.save(response)
        return response

    export_to_excel.short_description = "Export to Excel"

    actions = [export_to_excel]

class VendorAdmin(admin.ModelAdmin):
    list_display = ("name", "email", "phone", "city", "state")
    search_fields = list_display
    list_per_page = 25

    def export_to_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="vendors.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Vendors"

        # Write column headers
        headers = ["Name", "Email", "Phone", "City", "State"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_num, vendor in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=vendor.name)
            worksheet.cell(row=row_num, column=2, value=vendor.email)
            worksheet.cell(row=row_num, column=3, value=vendor.phone)
            worksheet.cell(row=row_num, column=4, value=vendor.city)
            worksheet.cell(row=row_num, column=5, value=vendor.state)

        # Auto-size columns
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save workbook to response
        workbook.save(response)
        return response

    export_to_excel.short_description = "Export to Excel"

    actions = [export_to_excel]

class StockInwardInline(admin.TabularInline):
    model = StockInward
    extra = 0


class StoreAdmin(admin.ModelAdmin):
    list_display = ("name", "description", "location", "city", "state")
    search_fields = list_display
    list_per_page = 25

    def export_to_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="stores.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Stores"

        # Write column headers
        headers = ["Name", "Description", "Location", "City", "State"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_num, store in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=store.name)
            worksheet.cell(row=row_num, column=2, value=store.description)
            worksheet.cell(row=row_num, column=3, value=store.location)
            worksheet.cell(row=row_num, column=4, value=store.city)
            worksheet.cell(row=row_num, column=5, value=store.state)

        # Auto-size columns
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save workbook to response
        workbook.save(response)
        return response

    export_to_excel.short_description = "Export to Excel"

    actions = [export_to_excel]
class ExpensesAdmin(admin.ModelAdmin):
    list_display = ("store", "description", "amount", "date")
    search_fields = list_display
    list_per_page = 25
    def export_to_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="expenses.csv"'

        writer = csv.writer(response)
        writer.writerow(["Store", "Description", "Amount", "Date"])

        for expense in queryset:
            writer.writerow([expense.store, expense.description, expense.amount, expense.date])

        return response

    export_to_csv.short_description = "Export to CSV"

    actions = [export_to_csv]

class StockInwardInline(admin.TabularInline):
    model = StockInward
    extra = 0


class StoreInline(admin.TabularInline):
    model = Store
    extra = 0


class StockItemsAdmin(admin.ModelAdmin):
    list_display = (
        "name",
        "description",
        "unit_price",
        "get_stock_inward",
    )
    search_fields = list_display
    list_per_page = 25

    def get_stock_inward(self, obj):
        return obj.stockinward_set.all().count()

    get_stock_inward.short_description = "Stock Inward Count"

    def export_to_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="stock_items.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Stock Items"

        # Write column headers
        headers = ["Name", "Description", "Unit Price", "Stock Inward Count"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_num, stock_item in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=stock_item.name)
            worksheet.cell(row=row_num, column=2, value=stock_item.description)
            worksheet.cell(row=row_num, column=3, value=stock_item.unit_price)
            worksheet.cell(row=row_num, column=4, value=stock_item.stockinward_set.count())

        # Auto-size columns
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save workbook to response
        workbook.save(response)
        return response

    export_to_excel.short_description = "Export to Excel"

    actions = [export_to_excel]


class StockInwardAdmin(ExportActionMixin, admin.ModelAdmin):
    def unit_price(self, obj):
        return format_html("<b>{}</b>", obj.stock_item_id.unit_price)

    readonly_fields = ("unit_price",)

    unit_price.short_description = "unit_price"

    list_display = [field.name for field in StockInward._meta.get_fields()]
    list_display.append("unit_price")

    list_filter = [field.name for field in StockInward._meta.fields]
    search_fields = list_display
    list_per_page = 25

    readonly_fields = [
        "invoice_number",
    ]
def total_price(self, obj):
        return format_html("{}", obj.stock_item_id.unit_price * obj.quantity)

class StockOutwardAdmin(ExportActionMixin, admin.ModelAdmin):
    def stock_inward_quantityce(self, obj):
        stock_item_id = obj.stock_item_id
        store_id = obj.store_id
        stock_inward = StockInward.objects.get(
            stock_item_id=stock_item_id, store_id=store_id
        )
        return stock_inward.quantity

    def unit_price(self, obj):
        return format_html("{}", obj.stock_item_id.unit_price)

    unit_price.short_description = "unit_price"

    def total_price(self, obj):
        return format_html("{}", obj.stock_item_id.unit_price * obj.quantity)

    def save_model(self, request, obj, form, change):
        # Check if the instance is being updated
        if change:
            original_outward = StockOutward.objects.get(pk=obj.pk)
            original_quantity = original_outward.quantity

            # Calculate the difference in quantity
            quantity_difference = obj.quantity - original_quantity

            print(
                "quantity_difference = obj.quantity - original_quantity",
                quantity_difference,
                obj.quantity,
                original_quantity,
            )
            # Get the corresponding StockInward instance
            stock_inward = StockInward.objects.get(
                stock_item_id=obj.stock_item_id, store_id=obj.store_id
            )

            stock_inward.quantity = stock_inward.quantity - quantity_difference
            stock_inward.save()

        super().save_model(request, obj, form, change)

    stock_inward_quantityce.short_description = "stock_inward_quantityce"

    list_display = [
        "store_id",
        "stock_item_id",
        "quantity",
        "customer_id",
        "recipient",
        "reason",
        "outward_date",
    ]
    list_display.insert(4, "unit_price")
    list_display.insert(5, "total_price")
    list_display.append("stock_inward_quantityce")
    search_fields = list_display
    list_per_page = 25
    list_filter = [field.name for field in StockOutward._meta.fields]
    readonly_fields = [
        "recipient",
        "unit_price",
        "total_price",
        "created_at",
        "updated_at",
    ]


class StockAdjustmentAdmin(admin.ModelAdmin):
    list_display = [field.name for field in StockAdjustment._meta.get_fields()]
    list_filter = [field.name for field in StockAdjustment._meta.fields]
    search_fields = list_display
    list_per_page = 25


admin.site.register(Client, ClientAdmin)
admin.site.register(Vendor, VendorAdmin)
admin.site.register(Store, StoreAdmin)
admin.site.register(Expenses, ExpensesAdmin)
admin.site.register(StockItems, StockItemsAdmin)
admin.site.register(StockInward, StockInwardAdmin)
admin.site.register(StockOutward, StockOutwardAdmin)
admin.site.register(StockAdjustment, StockAdjustmentAdmin)
